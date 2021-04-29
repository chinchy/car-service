from datetime import datetime
from flask import Blueprint, current_app, send_file
from openpyxl import Workbook
from os.path import basename, join
from sqlalchemy import case, func, Integer

from app.db import db
from .models import Car, Work, Service

bp = Blueprint('reports', __name__)


@bp.route('/all_works', methods=['GET', 'POST'])
def all_works():
    f_name = join(current_app.config['REPORT_FOLDER'],
                  "all_works_" + str(datetime.now().timestamp()).replace(".", "") + ".xlsx")

    wb = Workbook()
    ws = wb.active

    works = db.session.query(Work).order_by(Work.id).all()

    ws.cell(row=1, column=1).value = "Дата проведения работы"
    ws.cell(row=1, column=2).value = "Мастер"
    ws.cell(row=1, column=3).value = "Автомобиль"
    ws.cell(row=1, column=4).value = "Услуга"

    for i, row in enumerate(works, start=2):
        ws.cell(row=i, column=1).value = row.date_work
        ws.cell(row=i, column=2).value = str(row.master)
        ws.cell(row=i, column=3).value = str(row.car)
        ws.cell(row=i, column=4).value = str(row.service)

    wb.save(f_name)

    return send_file(
                    filename_or_fp=f_name,
                    as_attachment=True,
                    attachment_filename=basename(f_name),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@bp.route('/monthly_profit', methods=['GET', 'POST'])
def monthly_profit():
    months = {
        1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
        5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август', 9: 'Сентябрь',
        10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
    }

    f_name = join(current_app.config['REPORT_FOLDER'],
                  "month_profit_" + str(datetime.now().timestamp()).replace(".", "") + ".xlsx")

    wb = Workbook()
    ws = wb.active

    data = db.session.query(func.sum(case([(Car.is_foreign.is_(True), Service.cost_foreign),
                                           (Car.is_foreign.is_(False), Service.cost_our)],
                                          else_=0)),
                            func.cast(func.date_part('month', Work.date_work), Integer))\
        .select_from(Car).join(Work).join(Service)\
        .group_by(func.date_part('month', Work.date_work))\
        .order_by(func.date_part('month', Work.date_work)).all()

    ws.cell(row=1, column=1).value = "Месяц"
    ws.cell(row=1, column=2).value = "Прибыль"

    total_profit = 0
    for key, val in months.items():
        profit = list(filter(lambda x: x[1] == key, data))
        ws.cell(row=key+1, column=1).value = val
        ws.cell(row=key+1, column=2).value = profit[0][0] if profit else 0
        total_profit += ws.cell(row=key+1, column=2).value

    ws.cell(row=15, column=1).value = "ИТОГО:"
    ws.cell(row=15, column=2).value = total_profit

    wb.save(f_name)

    return send_file(
                    filename_or_fp=f_name,
                    as_attachment=True,
                    attachment_filename=basename(f_name),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
